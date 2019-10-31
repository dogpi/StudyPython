

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    # 打开文件
    file = load_workbook(filename=path)
    # print(file.get_sheet_names())
    sheets = file.sheetnames
    for sheetName in sheets:
        sheet = file[sheetName]
    # sheet = file.get_sheet_by_name(sheets[0])
    # # 最大行
    # print(sheet.max_row)
    # # 最大列
    # print(sheet.max_column)
    # # 表头名
    # print(sheet.title)
        for lineNum in range(1,sheet.max_row+1):
            lineList = []
            for columnNum in range(1,sheet.max_column+1):
                value = sheet.cell(row = lineNum,column=columnNum).value
                if value !=None:
                    lineList.append(value)
            print(lineList)
        print("*"*20)

path = r"D:\PythonCode\StudyCode\004_自动化办公\xlsx.xlsx"
readXlsxFile(path)